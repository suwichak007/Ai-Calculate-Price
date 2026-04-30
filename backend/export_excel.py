"""
export_excel.py — Excel report generator
  - generate_excel(result: dict) → bytes

  Requires: openpyxl
  pip install openpyxl
"""

import io
from datetime import datetime


def generate_excel(result: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Manday Cost"

    # ── Color palette ──────────────────────────────────────────
    DARK_BLUE   = "1E3A5F"
    MID_BLUE    = "2563EB"
    LIGHT_BLUE  = "EFF6FF"
    ACCENT      = "7C3AED"
    GREEN       = "059669"
    GREEN_LIGHT = "ECFDF5"
    GRAY_HDR    = "F1F5F9"
    GRAY_LINE   = "E2E8F0"
    WHITE       = "FFFFFF"

    def side(color=GRAY_LINE):
        return Side(style="thin", color=color)

    def border(all_color=GRAY_LINE):
        s = side(all_color)
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def bold_font(size=11, color="000000"):
        return Font(name="Arial", bold=True, size=size, color=color)

    def normal_font(size=11, color="000000"):
        return Font(name="Arial", size=size, color=color)

    # ── Column widths ──────────────────────────────────────────
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18

    # ── Header banner ──────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = "ใบประเมินค่าใช้จ่าย Manday"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color=WHITE)
    ws["A1"].fill = fill(DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"] = "Manday Cost Estimate"
    ws["A2"].font = Font(name="Arial", size=11, color="BFCFFD", italic=True)
    ws["A2"].fill = fill(DARK_BLUE)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # ── Meta info ──────────────────────────────────────────────
    def meta_row(r, label, value):
        ws.cell(r, 1, label).font = bold_font(10, "64748B")
        ws.cell(r, 2, value).font = normal_font(11)
        ws.cell(r, 1).fill = fill(GRAY_HDR)
        ws.cell(r, 2).fill = fill(WHITE)
        for c in range(1, 7):
            ws.cell(r, c).border = border(GRAY_LINE)
        ws.merge_cells(f"B{r}:F{r}")

    ws.merge_cells("A3:F3")
    ws["A3"].fill = fill(LIGHT_BLUE)
    ws.row_dimensions[3].height = 6

    meta_row(4, "ผู้จัดทำ / ผู้ขอ",     result["requester_name"])
    meta_row(5, "ชื่อโครงการ / ลูกค้า", result["project_name"])
    meta_row(6, "วันที่จัดทำ",           datetime.now().strftime("%d/%m/%Y %H:%M"))

    ws.merge_cells("A7:F7")
    ws["A7"].fill = fill(LIGHT_BLUE)
    ws.row_dimensions[7].height = 6

    # ── Section helpers ────────────────────────────────────────
    def section_header(r, title, color=MID_BLUE):
        ws.merge_cells(f"A{r}:F{r}")
        ws.cell(r, 1, f"  {title}").font = Font(name="Arial", bold=True, size=11, color=WHITE)
        ws.cell(r, 1).fill = fill(color)
        ws.cell(r, 1).alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 26

    def data_row(r, label, val, fmt="number", bg=WHITE):
        ws.cell(r, 1, label).font = normal_font(11)
        ws.cell(r, 1).fill = fill(bg)
        ws.cell(r, 1).border = border()
        ws.cell(r, 2).fill = fill(bg)
        ws.cell(r, 2).border = border()
        ws.merge_cells(f"B{r}:F{r}")
        c = ws.cell(r, 2, val)
        c.fill = fill(bg)
        c.border = border()
        c.alignment = Alignment(horizontal="right")
        if fmt == "number":
            c.font = Font(name="Arial Narrow", size=11)
            c.number_format = "#,##0"
        elif fmt == "pct":
            c.font = Font(name="Arial Narrow", size=11)
            c.number_format = '0.00"%"'
        elif fmt == "text":
            c.font = normal_font(11)
            c.alignment = Alignment(horizontal="left")

    def total_row(r, label, val, color=GREEN, bg=GREEN_LIGHT):
        ws.cell(r, 1, label).font = Font(name="Arial", bold=True, size=12, color=WHITE)
        ws.cell(r, 1).fill = fill(color)
        ws.cell(r, 1).border = border(color)
        ws.merge_cells(f"B{r}:F{r}")
        c = ws.cell(r, 2, val)
        c.font = Font(name="Arial Narrow", bold=True, size=13, color=color)
        c.fill = fill(bg)
        c.border = border(color)
        c.alignment = Alignment(horizontal="right")
        c.number_format = "#,##0"
        ws.row_dimensions[r].height = 28

    # ── Section 1: ข้อมูลพื้นฐาน ──────────────────────────────
    section_header(8, "📋  ข้อมูลพื้นฐาน")
    data_row(9,  "Manday รวมทุก Phase",       result["manday"],     "number", LIGHT_BLUE)
    data_row(10, "Markup %",                  result["markup_pct"], "number")

    # ── Section 2: ต้นทุนแยก 3 Phase ───────────────────────────
    section_header(12, "💰  ต้นทุนแยกตาม Phase")
    headers = ["Phase / หัวเรื่อง", "คน", "ครั้ง", "วัน/ครั้ง", "Rate", "ต้นทุน"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(13, c, h)
        cell.font = bold_font(10, WHITE)
        cell.fill = fill(DARK_BLUE)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border(DARK_BLUE)

    row = 14
    for phase in result.get("phase_costs", []):
        values = [phase["label"], "", "", "", "", phase["cost"]]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row, c, val)
            cell.fill = fill(LIGHT_BLUE)
            cell.border = border()
            cell.alignment = Alignment(horizontal="right" if c > 1 else "left")
            cell.font = Font(name="Arial Narrow" if c > 1 else "Arial", size=11, bold=True, color=MID_BLUE)
            if c > 1:
                cell.number_format = "#,##0"
        row += 1
        for item in phase.get("items", []):
            values = [item["title"], item["person"], item["times"], item["days"], item["rate"], item["cost"]]
            for c, val in enumerate(values, 1):
                cell = ws.cell(row, c, val)
                cell.fill = fill(WHITE)
                cell.border = border()
                cell.alignment = Alignment(horizontal="right" if c > 1 else "left")
                cell.font = Font(name="Arial Narrow" if c > 1 else "Arial", size=10)
                if c > 1:
                    cell.number_format = "#,##0"
            row += 1

    total_row(row, "รวมต้นทุน 3 Phase", result.get("subtotal_cost", 0), MID_BLUE, LIGHT_BLUE)
    ws.cell(row, 2).font = Font(name="Arial Narrow", bold=True, size=12, color=MID_BLUE)
    row += 2

    section_header(row, "📈  กำไรหลังรวม 3 Phase", GREEN)
    row += 1
    data_row(row, "Markup / กำไร %", result["markup_pct"], "number")
    row += 1
    data_row(row, f"กำไรหลังรวม 3 Phase ({result['markup_pct']}%)", result.get("profit", 0), "number", LIGHT_BLUE)
    ws.cell(row, 2).font = Font(name="Arial Narrow", bold=True, size=12, color=GREEN)

    # ── Section 3: ค่าเดินทาง ──────────────────────────────────
    td = result.get("travel_detail", {})
    travel_items = [
        (k, v) for k, v in [
            ("ค่าน้ำมัน",    td.get("fuel", 0)),
            ("ค่าโรงแรม",    td.get("hotel", 0)),
            ("เบี้ยเลี้ยง",   td.get("allowance", 0)),
            ("ค่าเครื่องบิน", td.get("flight", 0)),
            ("ค่าเช่ารถ",    td.get("rental", 0)),
            ("ค่า Taxi",     td.get("taxi", 0)),
            ("เบี้ยเดินทาง", td.get("travel_allow", 0)),
        ] if v > 0
    ]
    row += 1
    if travel_items:
        section_header(row, "✈️  รายละเอียดค่าเดินทาง (ต่อหน่วย)")
        row += 1
        for label, val in travel_items:
            data_row(row, label, val, "number")
            row += 1

    # ── Grand total ────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].fill = fill(LIGHT_BLUE)
    ws.row_dimensions[row].height = 8
    row += 1
    total_row(row, "🏆  ยอดรวมทั้งหมด", result["total"])

    # ── Footer ──────────────────────────────────────────────────
    footer_r = row + 2
    ws.merge_cells(f"A{footer_r}:F{footer_r}")
    ws[f"A{footer_r}"] = (
        f"จัดทำโดย: {result['requester_name']}   |   "
        f"วันที่: {datetime.now().strftime('%d/%m/%Y')}   |   Manday Cost Chatbot"
    )
    ws[f"A{footer_r}"].font = Font(name="Arial", size=9, color="94A3B8", italic=True)
    ws[f"A{footer_r}"].alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
