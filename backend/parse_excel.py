"""
parse_excel_upload.py — ใช้ LLM parse Excel แบบ flexible
  - อ่าน Excel → แปลงเป็น text table
  - ส่งให้ LLM วิเคราะห์และ extract ข้อมูลเป็น JSON
  - รองรับ format หลากหลาย ไม่ต้อง hardcode column/row

ส่งออก dict:
  {
    "project_name": "...",
    "customer":     "...",
    "phase_items":  [
        { "phase": "implement", "title": "...", "days": 2.5, "person": 1, "times": 1 },
        ...
    ],
    "warnings": ["..."]
  }
"""

import io
import json
import re

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── System prompt สำหรับ LLM parse Excel ──────────────────────────────────────

EXCEL_PARSE_SYSTEM = """You are a data extractor. Extract project plan data from Excel text.

CRITICAL: Respond ONLY with this EXACT JSON structure. No other structure allowed.

{
  "project_name": "string or null",
  "customer": "string or null",
  "engineers": "string or null",
  "phase_items": [
    {"phase": "prepare", "title": "string", "days": 0.5, "person": 1, "times": 1}
  ],
  "warnings": []
}

The "phase_items" array MUST be flat — no nesting, no sub-arrays, no "tasks" key.
Each task row becomes ONE object in phase_items directly.

PHASE MAPPING:
- "prepare"   = Phase I / Phase 1
- "implement" = Phase II / Phase III / Phase 2 / Phase 3 / Migrate / Install
- "service"   = Documentation / OJT / Support / MA

EXTRACTION RULES:
- Rows numbered "1.1", "1.2", "2.1", "2.2" etc → extract as phase_item
- Rows numbered "1", "2", "3" (no decimal) → section header, SKIP
- Rows starting with "-" → bullet detail, SKIP
- days: number only ("0.5 Day"→0.5, "2.5 Days"→2.5, "1 Day"→1)
- person: default 1
- times: default 1
- Sub-tasks like "Migrate WIN16-DC-AD1..." under "2.1" → each becomes its own phase_item

EXAMPLE OUTPUT for this input:
  1.1 | Site Survey | 0.5 Day
  1.2 | Fixing replication | 2.5 Days
  Migrate WIN16-DC-AD1 | 1 Day

→ phase_items: [
  {"phase":"prepare","title":"Site Survey","days":0.5,"person":1,"times":1},
  {"phase":"prepare","title":"Fixing replication issue on WIN-DC-AWS-1A","days":2.5,"person":1,"times":1},
  {"phase":"implement","title":"Migrate WIN16-DC-AD1 to Windows Server 2022","days":1,"person":1,"times":1}
]

CRITICAL: Return ONLY raw JSON. No markdown. No backticks. No explanation."""

def _excel_to_text(file_bytes: bytes) -> tuple:
    """
    อ่าน Excel -> text table format ที่ LLM อ่านได้ง่าย
    Returns (text_content, warnings)
    """
    warnings = []
    if not HAS_OPENPYXL:
        return "", ["openpyxl ไม่ได้ติดตั้ง — pip install openpyxl"]

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        # ลอง xlrd fallback สำหรับ .xls จริงๆ
        try:
            import pandas as pd
            import tempfile, os
            with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as f:
                f.write(file_bytes)
                tmp = f.name
            df = pd.read_excel(tmp, engine="xlrd")
            os.unlink(tmp)
            text = df.to_string(index=False)
            return text, ["ไฟล์เป็น .xls format — แปลงอัตโนมัติแล้ว"]
        except Exception:
            return "", [f"อ่านไฟล์ไม่ได้: {e}"]

    sheets_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = []

        max_col = ws.max_column or 6
        max_row = ws.max_row or 100

        # limit เพื่อไม่ให้ token เกิน
        row_limit = min(max_row, 200)
        col_limit = min(max_col, 8)

        for row in ws.iter_rows(min_row=1, max_row=row_limit, max_col=col_limit):
            cells = []
            for cell in row:
                val = cell.value
                if val is None:
                    cells.append("")
                elif isinstance(val, float):
                    cells.append(str(int(val)) if val == int(val) else str(round(val, 2)))
                else:
                    cells.append(str(val).strip().replace("\n", " "))

            # skip แถวว่างทั้งหมด
            if all(c == "" for c in cells):
                continue

            rows_data.append(" | ".join(cells))

        if rows_data:
            sheets_text.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows_data))

    if not sheets_text:
        warnings.append("ไม่พบข้อมูลในไฟล์ Excel")

    # รวมทุก sheet แล้วส่ง LLM (cap ~6000 chars)
    full_text = "\n\n".join(sheets_text)
    if len(full_text) > 60000:
        full_text = full_text[:60000] + "\n...[ข้อมูลถูกตัดเพราะยาวเกินไป]"
        warnings.append("ไฟล์มีข้อมูลมาก — อ่านได้เฉพาะ 200 แถวแรก")

    return full_text, warnings


def parse_project_excel(file_bytes: bytes, llm_instance=None) -> dict:
    """
    Main entry: parse bytes -> structured dict โดยใช้ LLM

    llm_instance: ส่ง LLM instance จาก router (ตัวเดียวกับที่ใช้ใน chat)
    """
    empty_result = {
        "project_name": "",
        "customer":     "",
        "engineers":    "",
        "phase_items":  [],
        "warnings":     [],
    }

    # Step 1: แปลง Excel -> text
    excel_text, read_warnings = _excel_to_text(file_bytes)
    if not excel_text:
        return {**empty_result, "warnings": read_warnings or ["อ่านไฟล์ไม่ได้"]}

    # Step 2: ส่งให้ LLM
    user_prompt = f"""วิเคราะห์ไฟล์ Excel Project Plan นี้และ extract ข้อมูลทั้งหมดเป็น JSON:

{excel_text}

กรุณา extract:
1. ชื่อโครงการ, ลูกค้า, วิศวกร (ถ้ามี)
2. หัวเรื่องย่อยทุกรายการ พร้อม phase (prepare/implement/service) และ Man Day
3. ใส่ warnings ถ้ามีข้อมูลไม่ครบหรือไม่แน่ใจ"""

    messages = [
        {"role": "system", "content": EXCEL_PARSE_SYSTEM},
        {"role": "user",   "content": user_prompt},
    ]

    raw_response = _call_llm(llm_instance, messages)

    # Step 3: parse JSON response
    result = _parse_llm_json(raw_response)
    if result is None:
        return {
            **empty_result,
            "warnings": read_warnings + ["LLM ตอบกลับไม่ถูกรูปแบบ — ลองอัพโหลดอีกครั้ง"],
        }

    # merge read_warnings
    result.setdefault("warnings", [])
    result["warnings"] = read_warnings + result["warnings"]

    # sanitize phase_items
    valid_phases = {"prepare", "implement", "service"}
    clean_items  = []
    for item in result.get("phase_items", []):
        if not isinstance(item, dict) or not item.get("title"):
            continue
        phase = str(item.get("phase", "implement")).lower().strip()
        if phase not in valid_phases:
            phase = "implement"
        clean_items.append({
            "phase":  phase,
            "title":  str(item["title"]).strip(),
            "days":   _safe_float(item.get("days"),   default=1.0),
            "person": _safe_float(item.get("person"), default=1.0),
            "times":  _safe_float(item.get("times"),  default=1.0),
        })

    result["phase_items"] = clean_items

    if not clean_items:
        result["warnings"].append(
            "ไม่พบรายการ task ในไฟล์ — ตรวจสอบว่ามีคอลัมน์ Task Detail และ Man Day"
        )

    return result


# ── LLM caller ────────────────────────────────────────────────────────────────

def _call_llm(llm_instance, messages: list) -> str:
    if llm_instance is not None:
        try:
            res = llm_instance.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=messages,  # ส่งทั้งก้อนเลย รวม system
                max_tokens=2000,
                temperature=0.1,
            )
            return res.choices[0].message.content.strip()
        except Exception as e:
            return json.dumps({"error": str(e)})
    return json.dumps({"error": "no llm instance"})


def _parse_llm_json(raw: str) -> dict | None:
    """parse JSON จาก LLM response — tolerant ต่อ markdown fences"""
    if not raw:
        return None
    text = re.sub(r"```(?:json)?\s*", "", raw).strip().rstrip("`").strip()

    start = text.find("{")
    end   = text.rfind("}") + 1
    if start == -1 or end == 0:
        return None

    try:
        return json.loads(text[start:end])
    except json.JSONDecodeError:
        try:
            fixed = re.sub(r",\s*([}\]])", r"\1", text[start:end])
            return json.loads(fixed)
        except Exception:
            return None


def _safe_float(val, default: float = 0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        m = re.search(r"(\d+\.?\d*)", str(val))
        return float(m.group(1)) if m else default